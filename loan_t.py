#!/usr/bin/env python
# coding: utf-8

# In[46]:


#importing all the required libraries for GUI
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox #for drop down in the forms
import tkinter as tk
#for storing the data in excel sheet
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import workbook
import pathlib 
 
root=Tk()

#Displaying the title 
root.title("Loan application form")
root.geometry("700x700+300+200")
root.resizable(False,False)
root.configure(bg="#326273")  
'''
file=pathlib.Path("Loan_data.xlsx")
if file.exists():
    pass
else:
    file=workbook()
    sheet=file.active
    sheet['A1']="Name"
    sheet['B1']='PhoneNumber'
    sheet['C1']='Age'
    sheet['D1']='Gender'
    sheet['E1']='Address'
    sheet['F1']='Monthly_Income'
    sheet['G1']='Members_in_family'
    
    file.save('Loan_data.xlsx')
    
def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=AdressEntry.get(1.0,END)
    income=MonthlyValue.get()
    members=MembersValue.get()
    
    file=openpyxl.load_workbook("Loan_data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row+1,value=contact)
    sheet.cell(column=3,row=sheet.max_row+1,value=age)
    sheet.cell(column=4,row=sheet.max_row+1,value=gender)
    sheet.cell(column=5,row=sheet.max_row+1,value=address)
    sheet.cell(column=6,row=sheet.max_row+1,value=income)
    sheet.cell(column=7,row=sheet.max_row+1,value=members)
    
    file.save(r'Loan_data.xlsx')
    '''
    
#for clearing all the data entered previously
def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    MonthlyValue.set('')
    MembersValue.set('')
    AdressEntry.delete(1.0,END)
    LoanValue.set('')
    IntrestValue.set('')
    MonthsValue.set('')
    
#function for calculating the monthly EMI based on loan amount,intrest rate and duration of repayment    
    
def payment():
    if LoanEntry.get() and IntrestEntry.get() and MonthsEntry.get():
        # Convert Entry Boxes to numbers
        months = int(MonthsEntry.get())
        rate = float(IntrestEntry.get())
        loan = int(LoanEntry.get())
        # Calculate The Loan
        # Monthly Interest Rate
        monthly_rate = rate / 100 / 12 
        # Get Our Payment
        payment = (monthly_rate / (1 - (1 + monthly_rate)**(-months))) * loan
        # Format Payment
        payment = f"{payment:,.2f}"

        # Output Payment to the screen
        payment_label.config(text=f"Monthly Payment: Rs.{payment}")

    else:
        payment_label.config(text="Please fill required Columns.")

#heading
Label(root,text="Loan Application form",font="arial 15",bg="#FA8072",fg="#fff").place(x=250,y=20)

#label
Label(root,text="Name",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=100)
Label(root,text="Contact No.",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=150)
Label(root,text="Age",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=200)
Label(root,text="Gender",font=23,bg="#48D1CC",fg="#fff").place(x=370,y=200)
Label(root,text="Address",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=250)
Label(root,text="Monthly Income",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=300)
Label(root,text="Members in Family",font=23,bg="#48D1CC",fg="#fff").place(x=350,y=300)
Label(root,text="Collateral",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=350)
Label(root,text="Loan Amount",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=450)
Label(root,text="Interest Rate",font=23,bg="#48D1CC",fg="#fff").place(x=350,y=450)
Label(root,text="Months",font=23,bg="#48D1CC",fg="#fff").place(x=50,y=500)

#Entry
nameValue=StringVar() #constructor with widget to accept values
contactValue=StringVar()
AgeValue=StringVar()
MonthlyValue=StringVar()
MembersValue=StringVar()
LoanValue=StringVar()
IntrestValue=StringVar()
MonthsValue=StringVar()

nameEntry=Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
contactEntry=Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
ageEntry=Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)
MonthlyEntry=Entry(root,textvariable=MonthlyValue,width=10,bd=2,font=20)
MembersEntry=Entry(root,textvariable=MembersValue,width=10,bd=2,font=20)
AdressEntry=Text(root,width=50,height=2,bd=4)
LoanEntry=Entry(root,textvariable=LoanValue,width=10,bd=2,font=20)
IntrestEntry=Entry(root,textvariable=IntrestValue,width=10,bd=2,font=20)
MonthsEntry=Entry(root,textvariable=MonthsValue,width=15,bd=2,font=20)
#gender
gender_combobox=Combobox(root,values=['Male','Female','Others'],font='arial 15',state='r',width=14)
gender_combobox.place(x=440,y=200)
#collateral
collateral_combobox=Combobox(root,values=['Salary','Gold','Vehicle','Property'],font='arial 15',state='r',width=14)
collateral_combobox.place(x=200,y=350)

nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
MonthlyEntry.place(x=200,y=300)
MembersEntry.place(x=520,y=300)
AdressEntry.place(x=200,y=240)
LoanEntry.place(x=200,y=450)
IntrestEntry.place(x=480,y=450)
MonthsEntry.place(x=200,y=500)


Button(root,text="Submit",bg="#FA8072",fg="white",width=15,height=2,command=payment).place(x=200,y=600)
Button(root,text="clear",bg="#FA8072",fg="white",width=15,height=2,command=clear).place(x=350,y=600)
Button(root,text="Exit",bg="#FA8072",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=500,y=600)

payment_label = Label(root, text="", bg="#326273",font='arial 15')
payment_label.pack(pady=50)


root.mainloop()


# In[ ]:





# In[ ]:




